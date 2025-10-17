from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from products.models import Product, StockSucursal
from sucursales.models import Sucursal
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
from django.utils.dateparse import parse_date

import os
from urllib.parse import urlparse, parse_qs

class Command(BaseCommand):
    help = "Import products from a CSV or XLSX file in a memory-friendly streaming way."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path or URL to CSV/XLSX file")
        parser.add_argument("--dry-run", action="store_true", help="Parse only, don't write DB")
        parser.add_argument("--batch", type=int, default=500, help="Batch size for bulk operations")

    def handle(self, path, dry_run=False, batch=500, **options):
        try:
            filetype = None
            plower = path.lower()
            if plower.endswith('.csv'):
                filetype = 'csv'
            elif plower.endswith('.xlsx'):
                filetype = 'xlsx'
            elif plower.startswith('http://') or plower.startswith('https://'):
                # Support Google Sheets export URLs like .../export?format=xlsx
                q = parse_qs(urlparse(path).query)
                fmt = (q.get('format') or [None])[0]
                if fmt in ('csv', 'xlsx'):
                    filetype = fmt

            if filetype == 'csv':
                created, updated = self._import_csv(path, dry_run, batch)
            elif filetype == 'xlsx':
                created, updated = self._import_xlsx(path, dry_run, batch)
            else:
                raise CommandError("Unsupported file type. Use .csv or .xlsx")
        except Exception as e:
            raise CommandError(str(e))

        self.stdout.write(self.style.SUCCESS(f"Imported. created={created}, updated={updated}"))

    def _safe_decimal(self, val):
        if val is None:
            return Decimal("0.00")
        try:
            return Decimal(str(val).strip().replace(",", "."))
        except (ValueError, TypeError, InvalidOperation):
            return Decimal("0.00")

    def _safe_int(self, val):
        try:
            if val is None:
                return 0
            s = str(val).strip().replace(",", "")
            if s == "":
                return 0
            return int(float(s))
        except Exception:
            return 0

    def _safe_bool(self, val, default=True):
        if val is None:
            return default
        s = str(val).strip().lower()
        if s in ("1", "true", "sí", "si", "yes", "y", "on"):  # Spanish variants included
            return True
        if s in ("0", "false", "no", "off", "n"):
            return False
        return default

    def _norm_date(self, fecha_raw):
        if not fecha_raw:
            return None
        if isinstance(fecha_raw, (datetime, date)):
            return fecha_raw.date() if isinstance(fecha_raw, datetime) else fecha_raw
        try:
            return parse_date(str(fecha_raw).split(" ")[0].strip())
        except Exception:
            return None

    def _process_rows(self, rows, header_map, dry_run=False, batch=500):
        def get_val(rv, header):
            idx = header_map.get(header)
            if idx is not None and idx < len(rv):
                return rv[idx]
            return None

        existing_map = {p.producto_id: p for p in Product.objects.filter(producto_id__isnull=False)}
        to_create = []
        to_update = []
        processed_codes = set()

        # Mapear sucursales por nombre (case-insensitive) e ID en texto
        sucursales = list(Sucursal.objects.all())
        suc_by_name = { (s.nombre or "").strip().lower(): s for s in sucursales }
        suc_by_id = { str(s.id): s for s in sucursales }

        # Detectar columnas de stock por sucursal. Soportamos prefijos:
        # "STOCK@Nombre", "STOCK:Nombre", "STOCK Nombre"
        stock_sucursal_cols = []  # lista de tuplas (header, sucursal)
        for h, i in header_map.items():
            if not h:
                continue
            hl = h.strip()
            base = None
            if hl.upper().startswith("STOCK@"):  # STOCK@Sucursal
                base = hl[6:].strip()
            elif hl.upper().startswith("STOCK:"):
                base = hl[6:].strip()
            elif hl.upper().startswith("STOCK "):
                base = hl[6:].strip()
            # Mapear base a sucursal
            if base:
                key = base.lower()
                suc = suc_by_name.get(key) or suc_by_id.get(base)
                if suc:
                    stock_sucursal_cols.append((h, suc))

        # Acumular cambios de stock por sucursal a aplicar tras flush de productos
        # Dict[(producto_id_code, sucursal_id)] = cantidad
        stocks_to_set = {}

        for row_values in rows:
            if not any(v for v in row_values if v is not None and str(v).strip() != ""):
                continue
            code = get_val(row_values, "CODIGO 1")
            if code is None or str(code).strip() == "":
                continue
            code = str(code).strip()
            if code in processed_codes:
                continue
            processed_codes.add(code)

            # Sucursal por fila (opcional)
            sucursal_val = get_val(row_values, "SUCURSAL")
            sucursal_obj = None
            if sucursal_val is not None:
                sstr = str(sucursal_val).strip()
                if sstr:
                    sucursal_obj = suc_by_name.get(sstr.lower()) or suc_by_id.get(sstr)

            defaults = {
                "nombre": str(get_val(row_values, "NOMBRE") or "").strip(),
                "descripcion": (str(get_val(row_values, "DESCRIPCION") or "").strip() or None),
                # Migración: interpretar CODIGO 2 como código de barras por defecto
                "codigo_barras": (str(get_val(row_values, "CODIGO DE BARRAS") or "") or str(get_val(row_values, "CODIGO 2") or "")).strip() or None,
                "codigo_alternativo": None,
                "fecha_ingreso_producto": self._norm_date(get_val(row_values, "FECHA DE INGRESO")),
                "precio_compra": self._safe_decimal(get_val(row_values, "PRECIO DE COMPRA")),
                "precio_venta": self._safe_decimal(get_val(row_values, "PRECIO DE VENTA")),
                "permitir_venta_sin_stock": self._safe_bool(get_val(row_values, "PERMITIR VENTA SIN STOCK"), default=True),
            }

            # Campos de cantidades globales (opcionales y de compatibilidad)
            cantidad_global = self._safe_int(get_val(row_values, "CANTIDAD"))
            stock_global = self._safe_int(get_val(row_values, "STOCK"))
            if cantidad_global:
                defaults["cantidad"] = cantidad_global
            if stock_global:
                defaults["stock"] = stock_global
            if sucursal_obj:
                defaults["sucursal"] = sucursal_obj

            if code in existing_map:
                prod = existing_map[code]
                if any(getattr(prod, k) != v for k, v in defaults.items()):
                    for k, v in defaults.items():
                        setattr(prod, k, v)
                    to_update.append(prod)
            else:
                to_create.append(Product(producto_id=code, **defaults))

            # Manejar stock por sucursal en fila
            # 1) Si hay columna STOCK y sucursal por fila → asignar stock a esa sucursal
            if stock_global and sucursal_obj:
                stocks_to_set[(code, sucursal_obj.id)] = stock_global

            # 2) Columnas dedicadas por sucursal detectadas
            for hdr, suc in stock_sucursal_cols:
                qty = self._safe_int(get_val(row_values, hdr))
                if qty:
                    stocks_to_set[(code, suc.id)] = qty

            # Flush periodically to keep memory low
            if not dry_run and (len(to_create) + len(to_update)) >= batch:
                self._flush(to_create, to_update, stocks_to_set, batch)

        if not dry_run:
            self._flush(to_create, to_update, stocks_to_set, batch)
            return len(to_create), len(to_update)
        return len(to_create), len(to_update)

    def _flush(self, to_create, to_update, stocks_to_set, batch):
        with transaction.atomic():
            if to_create:
                Product.objects.bulk_create(to_create, batch_size=batch)
                to_create.clear()
            if to_update:
                Product.objects.bulk_update(to_update, [
                    'nombre','descripcion','codigo_alternativo','codigo_barras','fecha_ingreso_producto','precio_compra','precio_venta','permitir_venta_sin_stock'
                ], batch_size=batch)
                to_update.clear()

            # Aplicar stocks por sucursal acumulados
            if stocks_to_set:
                # Mapear codigo -> producto
                codes = list({code for (code, _sid) in stocks_to_set.keys()})
                prod_map = {p.producto_id: p for p in Product.objects.filter(producto_id__in=codes)}
                # Upsert por par (producto, sucursal)
                for (code, suc_id), qty in list(stocks_to_set.items()):
                    prod = prod_map.get(code)
                    if not prod:
                        continue
                    ss, _created = StockSucursal.objects.get_or_create(producto=prod, sucursal_id=suc_id, defaults={"cantidad": max(0, qty)})
                    if not _created:
                        if ss.cantidad != qty:
                            ss.cantidad = max(0, qty)
                            ss.save()
                    # Remover del dict tras aplicar
                    stocks_to_set.pop((code, suc_id), None)

    def _import_csv(self, path, dry_run=False, batch=500):
        import csv
        import io
        import requests

        if path.startswith("http://") or path.startswith("https://"):
            resp = requests.get(path, stream=True, timeout=60)
            resp.raise_for_status()
            stream = io.TextIOWrapper(resp.raw, encoding='utf-8', newline='')
        else:
            stream = open(path, 'r', encoding='utf-8', newline='')

        with stream as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            header_map = {h.strip(): i for i, h in enumerate(headers) if h}
            minimal = ['NOMBRE', 'CODIGO 1', 'PRECIO DE COMPRA', 'PRECIO DE VENTA']
            missing = [h for h in minimal if h not in header_map]
            if missing:
                raise CommandError(f"Missing required headers: {', '.join(missing)}")
            return self._process_rows(reader, header_map, dry_run=dry_run, batch=batch)

    def _import_xlsx(self, path, dry_run=False, batch=500):
        from openpyxl import load_workbook
        import tempfile, requests
        import os

        # For URLs, download to a temp file to let openpyxl stream from disk
        cleanup = None
        if path.startswith("http://") or path.startswith("https://"):
            resp = requests.get(path, stream=True, timeout=120)
            resp.raise_for_status()
            fd, tmp = tempfile.mkstemp(suffix='.xlsx')
            os.close(fd)
            with open(tmp, 'wb') as out:
                for chunk in resp.iter_content(chunk_size=1024*1024):
                    if chunk:
                        out.write(chunk)
            path = tmp
            cleanup = tmp

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            sh = wb.active
            first_row = next(sh.iter_rows(min_row=1, max_row=1, values_only=True), None)
            headers = [str(v).strip() for v in (first_row or [])]
            header_map = {h: i for i, h in enumerate(headers) if h}
            minimal = ['NOMBRE', 'CODIGO 1', 'PRECIO DE COMPRA', 'PRECIO DE VENTA']
            missing = [h for h in minimal if h not in header_map]
            if missing:
                raise CommandError(f"Missing required headers: {', '.join(missing)}")

            def row_iter():
                for rv in sh.iter_rows(min_row=2, values_only=True):
                    yield rv

            created, updated = self._process_rows(row_iter(), header_map, dry_run=dry_run, batch=batch)
            try:
                wb.close()
            except Exception:
                pass
            return created, updated
        finally:
            if cleanup and os.path.exists(cleanup):
                try:
                    os.remove(cleanup)
                except Exception:
                    pass
