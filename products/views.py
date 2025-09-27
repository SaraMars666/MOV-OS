from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger 
from django.db.models import Q 
from .models import Product
from .forms import ProductForm
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from decimal import Decimal, InvalidOperation 
from datetime import datetime
from django.utils.dateparse import parse_date 

def product_management(request):
    """
    Vista para la gestión y listado de productos, incluyendo búsqueda, paginación y ordenamiento.
    """
    query = request.GET.get('search', '')
    
    # Obtener parámetros de ordenamiento
    sort_by = request.GET.get('sort_by', 'nombre')
    order = request.GET.get('order', 'asc')

    # Filtrar productos
    products = Product.objects.filter(
        Q(nombre__icontains=query) | 
        Q(producto_id__icontains=query) |
        Q(descripcion__icontains=query) |
        Q(proveedor__icontains=query)
    )

    # Lista de campos permitidos para ordenar
    allowed_sort_fields = {
        'nombre': 'nombre',
        'descripcion': 'descripcion',
        'codigo1': 'producto_id',
        'codigo2': 'codigo_alternativo',
        'proveedor': 'proveedor',
        'fecha_ingreso': 'fecha_ingreso_producto',
        'precio_compra': 'precio_compra',
        'precio_venta': 'precio_venta',
        'cantidad': 'cantidad',
        'stock': 'stock',
    }

    # Aplicar ordenamiento
    if sort_by in allowed_sort_fields:
        field_to_sort = allowed_sort_fields[sort_by]
        if order == 'desc':
            field_to_sort = '-' + field_to_sort
        products = products.order_by(field_to_sort)
    else:
        products = products.order_by('nombre') 

    total_products_count = products.count()
    per_page_options = [10, 25, 50, 100] 
    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10 
    
    if per_page not in per_page_options:
        per_page = 10 

    paginator = Paginator(products, per_page) 
    page = request.GET.get('page', 1) 
    
    try:
        products_page = paginator.page(page) 
    except PageNotAnInteger:
        products_page = paginator.page(1)
    except EmptyPage:
        products_page = paginator.page(paginator.num_pages)

    return render(request, 'products/product_management.html', {
        'products': products_page,
        'total_products_count': total_products_count,
        'search_query': query,
        'per_page': per_page, 
        'per_page_options': per_page_options, 
        'sort_by': sort_by, 
        'order': order, 
    })

def create_or_edit_product(request, product_id=None):
    """
    Vista para crear un nuevo producto o editar uno existente.
    """
    product = get_object_or_404(Product, id=product_id) if product_id else None
    form = ProductForm(request.POST or None, instance=product)
    title = 'Editar Producto' if product_id else 'Crear Producto'

    if request.method == 'POST' and form.is_valid():
        product_instance = form.save() 
        messages.success(request, 'Los cambios se guardaron con éxito.')
        if 'save_and_list' in request.POST:
            return redirect('product_management')
        return redirect('edit_product', product_instance.id)
    
    return render(request, 'products/product_form.html', {
        'form': form, 
        'title': title, 
        'product': product 
    })

def delete_product(request, product_id):
    """
    Vista para eliminar un producto.
    """
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Producto eliminado exitosamente.')
        return redirect('product_management')
    return render(request, 'products/delete_product.html', {'product': product})

def download_template(request):
    """
    Vista para descargar una plantilla Excel con los encabezados de los productos.
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Productos'
    
    # Encabezados simplificados
    headers = [
        'NOMBRE', 'DESCRIPCION', 'CODIGO 1', 'CODIGO 2', 'PROVEEDOR', 
        'FECHA DE INGRESO', 'PRECIO DE COMPRA', 'PRECIO DE VENTA'
    ]
    sheet.append(headers)

    workbook.save(response)
    return response

def upload_products(request):
    """
    Vista para subir productos desde un archivo Excel.
    """
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'No se subió ningún archivo.')
            return redirect('upload_products')

        file = request.FILES['file']

        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Por favor sube un archivo válido en formato Excel (.xlsx).')
            return redirect('upload_products')

        try:
            workbook = load_workbook(file, data_only=True)
            sheet = workbook.active

            header_row_values = [str(cell.value).strip() for cell in sheet[1]]
            header_map = {header: idx for idx, header in enumerate(header_row_values) if header}
            
            minimal_headers = ['NOMBRE', 'CODIGO 1', 'PRECIO DE COMPRA', 'PRECIO DE VENTA']
            
            for mh in minimal_headers:
                if mh not in header_map:
                    messages.error(request, f'La plantilla Excel no contiene el encabezado fundamental: "{mh}".')
                    return redirect('upload_products')

            products_processed = 0
            for row_idx, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(v for v in row_values if v is not None and str(v).strip() != ''):
                    messages.info(request, f'Fila {row_idx}: Fila completamente vacía detectada y saltada.')
                    continue

                try:
                    def get_val(header_name):
                        idx = header_map.get(header_name)
                        if idx is not None and idx < len(row_values):
                            return row_values[idx]
                        return None
                    
                    def to_decimal(val):
                        if val is None or (isinstance(val, str) and str(val).strip() == ''):
                            return Decimal('0.00')
                        try:
                            s_val = str(val).strip().replace(',', '.')
                            return Decimal(s_val)
                        except (ValueError, TypeError, InvalidOperation): 
                            return Decimal('0.00')

                    nombre = str(get_val('NOMBRE')).strip() if get_val('NOMBRE') is not None else None
                    descripcion = str(get_val('DESCRIPCION')).strip() if get_val('DESCRIPCION') is not None else None
                    producto_id_excel = str(get_val('CODIGO 1')).strip() if get_val('CODIGO 1') is not None else None
                    codigo_alternativo = str(get_val('CODIGO 2')).strip() if get_val('CODIGO 2') is not None else None
                    proveedor = str(get_val('PROVEEDOR')).strip() if get_val('PROVEEDOR') is not None else None
                    
                    fecha_ingreso_producto = None
                    fecha_raw = get_val('FECHA DE INGRESO')
                    if fecha_raw:
                        if isinstance(fecha_raw, (datetime, datetime.date)):
                            fecha_ingreso_producto = fecha_raw.date()
                        else:
                            try:
                                fecha_ingreso_producto = parse_date(str(fecha_raw).split(" ")[0].strip())
                            except (ValueError, TypeError):
                                messages.warning(request, f'Fila {row_idx}: Formato de fecha "{fecha_raw}" inválido. Se usará Nulo.')
                    
                    precio_compra = to_decimal(get_val('PRECIO DE COMPRA'))
                    precio_venta = to_decimal(get_val('PRECIO DE VENTA'))

                    # Validaciones
                    if not producto_id_excel:
                        messages.warning(request, f'Fila {row_idx}: El campo "CODIGO 1" está vacío. Esta fila será saltada.')
                        continue
                    
                    product_defaults = {
                        'nombre': nombre,
                        'descripcion': descripcion,
                        'codigo_alternativo': codigo_alternativo,
                        'proveedor': proveedor,
                        'fecha_ingreso_producto': fecha_ingreso_producto,
                        'precio_compra': precio_compra,
                        'precio_venta': precio_venta,
                        'permitir_venta_sin_stock': True,
                    }

                    product_obj, created = Product.objects.update_or_create(
                        producto_id=producto_id_excel,
                        defaults=product_defaults
                    )
                    products_processed += 1

                except Exception as e:
                    messages.error(request, f'Error al procesar fila {row_idx}: {str(e)}')
                    continue

            if products_processed > 0:
                messages.success(request, f'¡Se procesaron {products_processed} productos con éxito!')
            else:
                messages.warning(request, 'No se procesó ningún producto. Asegúrate de que el archivo no esté vacío.')
            
            return redirect('product_management')

        except Exception as e:
            messages.error(request, f'Error general procesando el archivo Excel: {str(e)}')
            return redirect('upload_products')

    return render(request, 'products/upload_products.html')

def delete_all_products(request):
    """
    Vista para eliminar todos los productos.
    """
    if request.method == 'POST':
        try:
            count, _ = Product.objects.all().delete()
            messages.success(request, f'¡Se eliminaron {count} productos exitosamente!')
            return redirect('product_management')
        except Exception as e:
            messages.error(request, f'Error al intentar eliminar productos: {str(e)}')
            return redirect('product_management')
    return render(request, 'products/delete_all_products_confirm.html')

def export_products_to_excel(request):
    """
    Vista para exportar todos los productos a un archivo Excel.
    Se utiliza el formateo definido en el modelo para los precios.
    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="export_productos.xlsx"'

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Productos'

    # Encabezados simplificados
    headers = [
        'NOMBRE', 'DESCRIPCION', 'CODIGO 1', 'CODIGO 2', 'PROVEEDOR', 
        'FECHA DE INGRESO', 'PRECIO DE COMPRA', 'PRECIO DE VENTA'
    ]
    sheet.append(headers)

    products = Product.objects.all().order_by('nombre')
    for product in products:
        row_data = [
            product.nombre,
            product.descripcion,
            product.producto_id,
            product.codigo_alternativo,
            product.proveedor,
            product.fecha_ingreso_producto,
            product.formatted_precio_compra,
            product.formatted_precio_venta,
        ]
        sheet.append(row_data)

    workbook.save(response)
    return response